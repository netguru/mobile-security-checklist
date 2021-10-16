"""Generator of Spreadsheet file as Mobile Security Checklist core tool"""
from string import ascii_uppercase
from os.path import dirname, abspath, join
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font,  PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

__headers = ['Code', 'Status', 'Category', 'Feature', 'Security requirement',
            'Solution in app / Comments', 'Last updated', 'Priority', 'Reference']
__statuses = ['TODO', 'DONE', 'Won\'t fix', 'Not applicable']
__priorities = ['Critical', 'High', 'Medium', 'Low']
__current_row = 1

def generate_spreadsheet(categories):
    """
    Generates content in 'mobile_security_checklist.xlsx' file for given Categories

    Parameters
    ----------
    categories: List[Category]
        list of Categories to fill '.xlsx' file for
    """

    global __current_row
    spreadsheet = load_workbook('checklist_base.xlsx')
    for category in categories[::-1]:
        __current_row = 1
        __generate_sheet(spreadsheet, category)
    spreadsheet.save(abspath(join(dirname(__file__), '..', 'mobile_security_checklist.xlsx')))

def  __generate_sheet(
    spreadsheet,
    category
):
    """
    Creates and fills new sheet for given Category

    Parameters
    ----------
    spreadsheet: Workbook
        model of the spreadsheet loaded from `.xlsx` file
    category: int
        category to generate new sheet for
    """

    new_sheet = __setup_sheet(
        spreadsheet,
        category
    )
    status_data_validation, priority_data_validation = __setup_data_validation(new_sheet)
    __setup_conditional_formatting(new_sheet)
    __generate_sheet_headers(new_sheet)
    __generate_requirement_rows(
        new_sheet,
        category,
        status_data_validation,
        priority_data_validation
    )
    __setup_data_filtering(new_sheet)

def __setup_sheet(
    spreadsheet,
    category
):
    """
    Creates new sheet with configured row freeze and column widths.

    Returns new Sheet object.

    Parameters
    ----------
    spreadsheet: Workbook
        model of the spreadsheet loaded from `.xlsx` file
    category: Category
        category to generate new sheet for
    """

    new_sheet = spreadsheet.create_sheet(category.name, 1)
    new_sheet.freeze_panes = f'{ascii_uppercase[0]}2'
    column_widths = [10, 20, 15, 45, 55, 30, 13, 15, 30]
    for column, width in zip(ascii_uppercase, column_widths):
        new_sheet.column_dimensions[column].width = width
    return new_sheet

def __setup_data_validation(sheet):
    """
    Add data validation to `Status` and `Priority` columns.

    Returns tuple of new DataValidation objects.

    Parameters
    ----------
    sheet: Sheet
        sheet to setup data validation for
    """

    status_data_validation = DataValidation(type='list', formula1=f'"{",".join(__statuses)}"')
    priority_data_validation = DataValidation(type='list', formula1=f'"{",".join(__priorities)}"')
    for data_validation in [status_data_validation, priority_data_validation]:
        sheet.add_data_validation(data_validation)
    return (status_data_validation, priority_data_validation)

def __setup_conditional_formatting(sheet):
    """
    Add conditional formatting to sheet rows depending on status

    Parameters
    ----------
    sheet: Sheet
        sheet to setup conditional formatting for
    """

    start_cell = f'{ascii_uppercase[0]}2'
    end_cell = f'{ascii_uppercase[len(__headers)-1]}980'
    colors = ['00FFFFFF', '00B7E1CD', '00FCE8B2', '00B7B7B7']
    conditional_formattings = {}
    for status, color in zip(__statuses, colors):
        conditional_formattings[status] = {
            "range": f'{start_cell}:{end_cell}',
            "formula": f'${ascii_uppercase[1]}2="{status}"',
            "style": DifferentialStyle(fill=PatternFill(bgColor=color, fill_type='solid'))
        }
    for conditional_formatting in conditional_formattings.values():
        sheet.conditional_formatting.add(
            conditional_formatting['range'],
            Rule(
                type='expression',
                formula=[conditional_formatting['formula']],
                dxf=conditional_formatting['style']
            )
        )

def __generate_sheet_headers(sheet):
    """
    Fill first sheet row with headers

    Parameters
    ----------
    sheet: Sheet
        sheet to generate headers for
    """

    global __current_row
    for column, header in zip(ascii_uppercase, __headers):
        sheet[f'{column}{__current_row}'] = header
        sheet[f'{column}{__current_row}'].font = Font(bold=True)
        sheet[f'{column}{__current_row}'].border = Border()
    __current_row += 1

def __generate_requirement_rows(
    sheet,
    category,
    status_data_validation,
    priority_data_validation
):
    """
    Fills sheet rows with requirements data for given category

    Parameters
    ----------
    sheet: Sheet
        sheet to fill rows in
    category: int
        category to generate rows for
    status_data_validation: DataValidation
        data validation model for `Status` column
    priority_data_validation: DataValidation
        data validation model for `Priority` column
    """

    global __current_row
    for group in category.groups:
        for requirement in group.requirements:
            __fill_row_for_requirement(
                sheet,
                category,
                group,
                requirement
            )
            __setup_date_format(sheet)
            __setup_reference_style(
                sheet,
                requirement
            )
            __setup_row_data_validation(
                sheet,
                status_data_validation,
                priority_data_validation
            )
            __current_row += 1

def __generate_requirement_data(
    category,
    group,
    requirement
):
    """
    Generates data dict for given category, requirements group and requirement.

    Returns new dict object.

    Parameters
    ----------
    category: Category
        category of the requirement
    group: RequirementsGroup
        requirements group of the requirement
    requirement: Requirement
        requirement to generate data for
    """

    should_add_header = len(category.groups) > 1
    status_cell = f'{ascii_uppercase[1]}{__current_row}'
    date_cell = f'{ascii_uppercase[5]}{__current_row}'
    data = [
        f'{category.code}{group.code if should_add_header else ""}.{requirement.uuid}',
            __statuses[0],
            group.name,
            requirement.feature.replace("<br>","\n"),
            requirement.description.replace("<br>","\n"),
            '',
            f'=IFERROR({status_cell}+{date_cell}{__current_row}+"x",TODAY())',
            requirement.priority.title()
    ]
    return data

def __fill_row_for_requirement(
    sheet,
    category,
    group,
    requirement
):
    """
    Fills new row with data for given requiremnt

    Parameters
    ----------
    sheet: Sheet
        sheet to fill with data
    category: Category
        category of the requirement
    group: RequirementsGroup
        requirements group of the requirement
    requirement: Requirement
        requirement to fill row with data for
    """

    data = __generate_requirement_data(category, group, requirement)
    for column, value in zip(ascii_uppercase, data):
        sheet[f'{column}{__current_row}'] = value
        sheet[f'{column}{__current_row}'].alignment = Alignment(wrap_text=True)

def __setup_date_format(sheet):
    """
    Apply proper style and format for 'Last updated' column

    Parameters
    ----------
    sheet: Sheet
        sheet to setup conditional formatting for
    """

    last_updated_col = f'{ascii_uppercase[6]}{__current_row}'
    sheet[last_updated_col].number_format = 'DD/MM/YYYY'
    sheet[last_updated_col].alignment = Alignment(horizontal="right", wrap_text=True)

def __setup_reference_style(
    sheet,
    requirement
):
    """
    Apply proper style and url for 'Reference' column

    Parameters
    ----------
    sheet: Sheet
        sheet to setup reference stlye for
    requirement: Requirement
        requirement to setup reference link for
    """

    repo_url = 'https://github.com/netguru/mobile-security-checklist/tree/master/'
    reference_col = f'{ascii_uppercase[8]}{__current_row}'
    hyperlink = f'{repo_url}{requirement.reference.replace("../", "")}'
    sheet[reference_col].hyperlink = hyperlink
    sheet[reference_col].value  = 'Handbook'
    sheet[reference_col].style = 'Hyperlink'

def __setup_row_data_validation(
    sheet,
    status_data_validation,
    priority_data_validation
):
    """
    Apply proper style and url for 'Reference' column

    Parameters
    ----------
    sheet: Sheet
        sheet to setup data validation for
    status_data_validation: DataValidation
        data validation model for `Status` column
    priority_data_validation: DataValidation
        data validation model for `Priority` column
    """

    status_data_validation.add(sheet[f'{ascii_uppercase[1]}{__current_row}'])
    priority_data_validation.add(sheet[f'{ascii_uppercase[7]}{__current_row}'])

def __setup_data_filtering(sheet):
    """
    Apply data filters to `Category` column

    Parameters
    ----------
    sheet: Sheet
        sheet to setup data filtering for
    """

    sheet.auto_filter.ref = f'{ascii_uppercase[2]}1:{ascii_uppercase[2]}{__current_row-1}'